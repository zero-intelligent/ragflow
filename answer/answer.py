import glob
import json
import os
import re
from concurrent.futures import ThreadPoolExecutor

import pandas as pd
import requests
from openpyxl.reader.excel import load_workbook

REQ_AUTHORIZATION = 'IjRlMDJjN2I2OWNlOTExZWY4NzgzMDI0MmFjMTIwMDA2Ig.ZyyF_A.EMdCm4ZCG4ZV4VWv8lbr2WesV2A'
DIALOG_ID = '8e31b0009d7411ef985d0242ac120006'  # '8e70f44885e311efb24b0242ac120006'
REQ_CONTENT_TYPE = 'application/json;charset=UTF-8'

RESULT_PATH = './数据-result.xlsx'
ORIGIN_PATH = './数据.xlsx'
TEST_LINES = 20000
START_ROW = 0
CONCURRENCY_NUM = 10


# 替换为你的Excel文件路径
def readExcel(file_path):
    return pd.read_excel(file_path)


def getRef(cid):
    url = 'http://39.101.69.172:18080/v1/conversation/get'
    params = {"conversation_id": cid}

    # 自定义请求头
    headers = {
        'Authorization': REQ_AUTHORIZATION
    }

    response = requests.get(url, params=params, headers=headers)

    if response.status_code == 200:
        data = response.json()  # 如果返回的是JSON格式的数据
        try:
            if 'reference' in data['data'] and data['data']['reference']:
                refObjs = [docAggs["doc_aggs"] for docAggs in data['data']['reference']]
                refs = [item['doc_name'] for sub in refObjs for item in sub]
                return refs
        except Exception as e:
            print('ref 获取失败!', data, e)
    else:
        print(f"ref 请求失败: {response}")
        return None;


def genConversationId():
    url = 'http://39.101.69.172:18080/v1/conversation/set'
    payload = {"dialog_id": DIALOG_ID, "name": "你好",
               "message": [{"role": "assistant", "content": "你好"}]}

    # 自定义请求头
    headers = {
        'Content-Type': REQ_CONTENT_TYPE,
        'Authorization': REQ_AUTHORIZATION
    }

    response = requests.post(url, json=payload, headers=headers)

    if response.status_code == 200:
        data = response.json()  # 如果返回的是JSON格式的数据
        if data['data'] and data['data']['id']:
            return data['data']['id']
    else:
        print(f"请求失败: {response}")
        return None;


def req(questionText: str, cId):
    url = 'http://39.101.69.172:18080/v1/conversation/completion'

    payload = {
        "conversation_id": f"{cId}",
        "messages": [
            {"content": "你好！ 我是你的助理，有什么可以帮到你的吗？", "role": "assistant"},
            {
                "content": questionText,
                "role": "user",
                "doc_ids": []
            }
        ]
    }

    # 自定义请求头
    headers = {
        'Content-Type': REQ_CONTENT_TYPE,
        'Authorization': REQ_AUTHORIZATION
    }
    response = requests.post(url, json=payload, headers=headers)
    last_answer = None
    if response.status_code == 200:
        for line in response.iter_lines():
            if line:  # 确保行不为空
                # 处理每一行事件
                line = line.decode('utf-8')[5:]
                try:
                    json_data = json.loads(line)['data']
                    if isinstance(json_data, dict):
                        last_answer = json_data['answer']  # 更新最后一个answer
                except json.JSONDecodeError as e:
                    print("JSON解析错误", e)

    else:
        print(f"请求失败，状态码: {response.status_code}")

    return last_answer


def write2Excel(originRow, rag, refs, ws, dir='./', fileName="result.xlsx"):
    ws.append(
        [originRow.iloc[0], originRow.iloc[1], originRow.iloc[2], originRow.iloc[3], originRow.iloc[4], rag,
         ' '.join([] if refs is None else refs)])


def splitExcelFile(fileNum):
    df = readExcel(ORIGIN_PATH)

    # 计算每个文件应包含的行数 地板除法
    rows_per_file = len(df) // fileNum
    remainder = len(df) % fileNum
    start_row=0
    end_row = 0
    # 分割 DataFrame 并保存为不同的 Excel 文件
    for i in range(fileNum):
        start_row = end_row;
        # 如果是最后一个文件，包含剩余所有行
        if i == fileNum - 1:
            df_chunk = df[start_row:]
        else:
            extra = 1 if i < remainder else 0
            end_row = start_row + rows_per_file + extra
            df_chunk = df[start_row:end_row]

        print(f"###########{start_row}, {end_row}")
        resultName = os.path.abspath(os.path.join(os.path.dirname(ORIGIN_PATH),
                                                  os.path.basename(ORIGIN_PATH).replace('.xlsx', f'_{i + 1}.xlsx')))
        df_chunk.to_excel(resultName, index=False)


def answer(originPath, resultPath):
    rows = readExcel(originPath).iterrows()
    if not os.path.exists(resultPath):
        df = pd.DataFrame()
        df.to_excel(resultPath, index=False)

    wb = load_workbook(resultPath)
    # wb = Workbook()
    ws = wb.active

    df = pd.read_excel(resultPath)

    # 检查是否有标题
    if df.empty or df.columns.empty:
        columns = ['疾病', '类型', '描述', '病人', '医生', 'RAG回复结果', '参考书籍']
        ws.append(columns)

    ct = 0;

    for index, row in rows:
        cId = genConversationId()

        if index >= START_ROW and ct < TEST_LINES:
            ct += 1
            question = row.iloc[3]
            ans = req(question, cId)
            refs = getRef(cId)
            print(f'############# {index} ###################')
            print(question)
            print('---------------------------')
            print(ans)
            print('---------------------------')
            print(row.iloc[4])
            print('---------------------------')
            print(refs)
            print('################################')

            write2Excel(row, ans, refs, ws)
            wb.save(resultPath)


def mergeFiles(resultFile):
    # 定义文件的通配符路径，n 是变量，这里使用 * 来匹配数字

    file_pattern = os.path.abspath(os.path.join(os.path.dirname(resultFile),
                                                os.path.basename(resultFile).replace('.xlsx', f'_*.xlsx')))
    # 使用 glob 获取所有匹配的文件名
    files = glob.glob(file_pattern)

    def extract_number(file_name):
        match = re.search(r'result_(\d+)', file_name)
        if match:
            return int(match.group(1))  # 提取并返回数字部分
        return 0

    # 对文件列表按 n 进行排序
    files.sort(key=lambda x: extract_number(x))

    # 创建一个空的 DataFrame 来存放合并后的数据
    combined_df = pd.DataFrame()

    # 遍历所有文件，将它们的数据读取并合并
    for file in files:
        df = pd.read_excel(file)  # 读取当前文件
        combined_df = pd.concat([combined_df, df], ignore_index=True)  # 合并数据

    # 保存合并后的数据到新的 Excel 文件
    combined_df.to_excel(resultFile, index=False)
    print(f"成功合并 {len(files)} 个文件，合并后的数据已保存为 {resultFile}")


def clearFiles(filePath):
    # 构建匹配的文件路径模式
    file_pattern = os.path.abspath(os.path.join(os.path.dirname(filePath),
                                                os.path.basename(filePath).replace('.xlsx', f'_*.xlsx')))
    # 使用 glob 获取所有匹配的文件名
    files_to_delete = glob.glob(file_pattern)

    # 删除所有匹配的文件
    for file in files_to_delete:
        try:
            os.remove(file)  # 删除文件
            print(f"已删除文件: {file}")
        except Exception as e:
            print(f"删除文件 {file} 时发生错误: {e}")

    print(f"已成功删除 {len(files_to_delete)} 个文件。")


def main():
    clearFiles(RESULT_PATH)
    clearFiles(ORIGIN_PATH)
    splitExcelFile(CONCURRENCY_NUM)

    with ThreadPoolExecutor(max_workers=CONCURRENCY_NUM) as executor:

        futures = []
        for i in range(CONCURRENCY_NUM):
            originName = os.path.abspath(os.path.join(os.path.dirname(ORIGIN_PATH),
                                                      os.path.basename(ORIGIN_PATH).replace('.xlsx', f'_{i + 1}.xlsx')))
            resultName = os.path.abspath(os.path.join(os.path.dirname(ORIGIN_PATH),
                                                      os.path.basename(RESULT_PATH).replace('.xlsx', f'_{i + 1}.xlsx')))

            futures.append(executor.submit(answer, originName, resultName))
        for future in futures:
            try:
                print(f"结果: {future.result()}")
            except Exception as e:
                print(f"任务引发异常: {e}")

    mergeFiles(RESULT_PATH)
    clearFiles(RESULT_PATH)
    clearFiles(ORIGIN_PATH)


if __name__ == '__main__':
    main()
